#!/usr/bin/env python3
"""Export a read-only earnings-calendar and financial-data Web Snapshot."""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import tempfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
RESEARCHOS_ROOT = PROJECT_DIR.parent
SNAPSHOT_PATH = PROJECT_DIR / "snapshot/data_platform_snapshot.json"
PUBLIC_DIR = PROJECT_DIR / "public"
CALENDAR_SUPPLEMENTS_PATH = PROJECT_DIR / "frontend_data/earnings_calendar_supplements.json"
CALENDAR_COVERAGE_REVIEWS_PATH = PROJECT_DIR / "frontend_data/calendar_coverage_reviews.json"
SNAPSHOT_SCHEMA = "ResearchOS-DataPlatformSnapshot-0.1"
BLOCKED_BROWSER_TERMS = (
    'data-panel="conclusions"',
    'data-panel="statements"',
    'data-panel="missing"',
    "研究结论",
    "管理层表态",
    "研究缺口",
    "research_output_id",
    "finding_id",
    "statement_id",
)


class SnapshotError(RuntimeError):
    """A fail-closed snapshot export error."""


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def atomic_write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=path.name + ".", dir=path.parent)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def read_sheet(relative_path: str, sheet_name: str) -> list[dict[str, Any]]:
    path = RESEARCHOS_ROOT / relative_path
    if not path.is_file():
        raise SnapshotError(f"正式数据来源缺失：{relative_path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SnapshotError(f"正式数据来源缺少工作表：{relative_path} / {sheet_name}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        try:
            headers = [str(value or "") for value in next(rows)]
        except StopIteration as exc:
            raise SnapshotError(f"正式数据工作表为空：{relative_path} / {sheet_name}") from exc
        return [
            {header: cell_value(value) for header, value in zip(headers, row)}
            for row in rows
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()


def e(value: Any) -> str:
    return html.escape(str(value if value is not None else ""), quote=True)


def source_manifest() -> dict[str, dict[str, Any]]:
    relative_paths = (
        "00_系统/关注公司.xlsx",
        "00_系统/配置/MetricDefinitionRegistry_v0.1.json",
        "01_财报日历/财报日历.xlsx",
        "02_原始财报/财报材料索引.xlsx",
        "04_公司数据库/财报历史数据库.xlsx",
    )
    manifest: dict[str, dict[str, Any]] = {}
    for relative in relative_paths:
        path = RESEARCHOS_ROOT / relative
        if not path.is_file():
            raise SnapshotError(f"数据平台正式来源缺失：{relative}")
        manifest[relative] = {"sha256": sha256_file(path), "size_bytes": path.stat().st_size}
    return manifest


def frontend_input_manifest() -> dict[str, dict[str, Any]]:
    paths = (CALENDAR_SUPPLEMENTS_PATH, CALENDAR_COVERAGE_REVIEWS_PATH)
    manifest = {}
    for path in paths:
        if not path.is_file():
            raise SnapshotError(f"前端数据规则文件缺失：{path.relative_to(PROJECT_DIR)}")
        relative = str(path.relative_to(PROJECT_DIR))
        manifest[relative] = {"sha256": sha256_file(path), "size_bytes": path.stat().st_size}
    return manifest


def verify_file_manifest(manifest: Any, root: Path, label: str) -> None:
    if not isinstance(manifest, dict) or not manifest:
        raise SnapshotError(f"{label} manifest 无效")
    for relative, expected in manifest.items():
        path = root / relative
        if not path.is_file():
            raise SnapshotError(f"{label}文件缺失：{relative}")
        if expected.get("sha256") != sha256_file(path) or expected.get("size_bytes") != path.stat().st_size:
            raise SnapshotError(f"{label}已变化，必须重新生成快照：{relative}")


def current_row(row: dict[str, Any]) -> bool:
    return str(row.get("is_current") or "").strip().lower() in {"是", "true", "1", "yes"}


def normalized_unit(row: dict[str, Any]) -> str:
    unit = str(row.get("normalized_unit") or row.get("raw_unit") or "")
    if unit == "million_currency":
        return f'{row.get("currency") or ""} 百万'.strip()
    if unit == "percent":
        return "%"
    if unit == "vehicles":
        return "辆"
    return unit or "未注明"


def scope_part(row: dict[str, Any], id_key: str, name_key: str) -> str | None:
    identity = str(row.get(id_key) or "consolidated")
    if identity == "consolidated":
        return None
    return str(row.get(name_key) or identity)


def series_label(row: dict[str, Any]) -> str:
    dimensions = [
        scope_part(row, "business_id", "业务名称"),
        scope_part(row, "geography_id", "地区名称"),
        scope_part(row, "product_id", "产品名称"),
    ]
    parts = [part for part in dimensions if part]
    parts.append(str(row.get("指标名称") or row.get("metric_id") or "未命名指标"))
    return " · ".join(parts)


def series_identity(row: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        str(row.get(key) or "")
        for key in (
            "metric_id",
            "business_id",
            "geography_id",
            "product_id",
            "basis",
            "currency",
            "normalized_unit",
            "period_type",
            "metric_definition_id",
            "measurement_basis",
        )
    )


PERIOD_TYPE_LABELS = {
    "quarter": "单季",
    "ytd_6m": "上半年累计",
    "ytd_9m": "前三季度累计",
    "fy": "全年",
    "point_in_time": "期末",
}

BASIS_LABELS = {
    "PRC_GAAP": "中国会计准则",
    "PRC_GAAP_attributable_to_parent": "中国会计准则归母",
    "HKFRS_attributable_to_parent": "HKFRS归母",
    "company_operating_metric": "公司运营口径",
}


def disambiguate_series_labels(series: list[dict[str, Any]]) -> None:
    by_label: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in series:
        by_label[item["label"]].append(item)
    for base_label, duplicates in by_label.items():
        if len(duplicates) < 2:
            continue
        for item in duplicates:
            qualifiers = [
                PERIOD_TYPE_LABELS.get(str(item.get("period_type") or ""), str(item.get("period_type") or "未注明期间")),
                BASIS_LABELS.get(str(item.get("basis") or ""), str(item.get("basis") or "未注明口径")),
            ]
            item["label"] = f'{base_label}（{" · ".join(qualifiers)}）'

    by_qualified_label: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in series:
        by_qualified_label[item["label"]].append(item)
    for qualified_label, duplicates in by_qualified_label.items():
        if len(duplicates) < 2:
            continue
        for item in duplicates:
            periods = [str(record.get("period") or "") for record in item["records"] if record.get("period")]
            period_range = periods[0] if len(periods) == 1 else f"{periods[0]}–{periods[-1]}"
            item["label"] = f"{qualified_label[:-1]} · {period_range}）"


def parse_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def percent_change(current: Any, previous: Any) -> dict[str, Any]:
    try:
        current_number = float(current)
        previous_number = float(previous)
    except (TypeError, ValueError):
        return {"value": None, "reason_code": "non_numeric_input"}
    if previous_number == 0:
        return {"value": None, "reason_code": "zero_denominator"}
    return {
        "value": (current_number / previous_number - 1) * 100,
        "reason_code": "frontend_program_calculation",
        "warnings": [],
    }


def comparison_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for index, record in enumerate(records):
        item = dict(record)
        current_date = parse_date(record.get("period_end_date"))
        prior_rows = records[:index]
        qoq_source = None
        yoy_source = None
        if current_date:
            dated = [(candidate, parse_date(candidate.get("period_end_date"))) for candidate in prior_rows]
            dated = [(candidate, candidate_date) for candidate, candidate_date in dated if candidate_date]
            qoq_candidates = [
                (candidate, (current_date - candidate_date).days)
                for candidate, candidate_date in dated
                if 45 <= (current_date - candidate_date).days <= 160
            ]
            yoy_candidates = [
                (candidate, abs((current_date - candidate_date).days - 365))
                for candidate, candidate_date in dated
                if 300 <= (current_date - candidate_date).days <= 430
            ]
            if qoq_candidates:
                qoq_source = min(qoq_candidates, key=lambda pair: pair[1])[0]
            if yoy_candidates:
                yoy_source = min(yoy_candidates, key=lambda pair: pair[1])[0]
        period_type = str(record.get("period_type") or "")
        item["qoq"] = (
            percent_change(record.get("value"), qoq_source.get("value"))
            if qoq_source and period_type in {"quarter", "point_in_time"}
            else {"value": None, "reason_code": "not_comparable_or_missing_prior_period"}
        )
        item["yoy"] = (
            percent_change(record.get("value"), yoy_source.get("value"))
            if yoy_source
            else {"value": None, "reason_code": "missing_prior_year_period"}
        )
        result.append(item)
    return result


def load_registry() -> dict[str, dict[str, Any]]:
    path = RESEARCHOS_ROOT / "00_系统/配置/MetricDefinitionRegistry_v0.1.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("mode") != "production":
        raise SnapshotError("Metric Definition Registry 不是 production")
    return {
        str(row["metric_definition_id"]): row
        for row in payload.get("definitions", [])
        if row.get("metric_definition_id")
    }


def build_series(rows: list[dict[str, Any]], registry: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("record_id") and current_row(row) and row.get("normalized_value") is not None:
            grouped[series_identity(row)].append(row)
    series = []
    for identity, candidates in grouped.items():
        candidates.sort(key=lambda row: (str(row.get("period_end_date") or ""), str(row.get("财报期") or "")))
        selected = candidates[-8:]
        first = selected[-1]
        definition_id = str(first.get("metric_definition_id") or "")
        definition = registry.get(definition_id, {})
        records = comparison_records(
            [
                {
                    "period": row.get("财报期"),
                    "period_type": row.get("period_type"),
                    "period_start_date": row.get("period_start_date"),
                    "period_end_date": row.get("period_end_date"),
                    "value": row.get("normalized_value"),
                    "unit": normalized_unit(row),
                    "currency": row.get("currency"),
                    "basis": row.get("basis"),
                    "measurement_basis": row.get("measurement_basis"),
                    "data_nature": row.get("数据性质"),
                    "data_status": row.get("数据状态"),
                    "record_id": row.get("record_id"),
                    "material_id": row.get("primary_material_id"),
                    "source_location": row.get("source_location"),
                    "formula_id": row.get("formula_id"),
                    "formula": row.get("计算公式"),
                    "input_record_ids": row.get("input_record_ids"),
                    "definition_id": definition_id,
                }
                for row in selected
            ]
        )
        series_seed = "|".join(identity).encode("utf-8")
        definition_view = {
            "name": definition.get("definition_name") or first.get("指标名称") or first.get("metric_id"),
            "metric_definition_id": definition_id or first.get("metric_id"),
            "description": definition.get("definition_description") or "按正式历史数据库当前版本展示。",
            "scope_relationship": definition.get("scope_relationship") or "按独立 business / geography / product 维度保存。",
            "verification_status": definition.get("verification_status") or first.get("数据状态"),
        }
        base_label = series_label(first)
        series.append(
            {
                "series_id": "series_" + sha256_bytes(series_seed)[:16],
                "base_label": base_label,
                "label": base_label,
                "unit": normalized_unit(first),
                "basis": first.get("basis") or "未注明",
                "period_type": first.get("period_type"),
                "records": records,
                "definitions": [definition_view],
            }
        )
    disambiguate_series_labels(series)
    return sorted(series, key=lambda item: (item["label"], item["basis"], item["period_type"] or ""))


def verify_series_labels(company_id: str, series: list[dict[str, Any]]) -> None:
    labels = [str(item.get("label") or "") for item in series]
    if any(not label for label in labels):
        raise SnapshotError(f"趋势指标缺少显示名称：{company_id}")
    if len(labels) != len(set(labels)):
        raise SnapshotError(f"趋势指标显示名称重复：{company_id}")
    series_ids = [str(item.get("series_id") or "") for item in series]
    if any(not series_id for series_id in series_ids) or len(series_ids) != len(set(series_ids)):
        raise SnapshotError(f"趋势指标 series_id 缺失或重复：{company_id}")

    by_base_label: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in series:
        by_base_label[str(item.get("base_label") or item.get("label") or "")].append(item)
    for base_label, duplicates in by_base_label.items():
        if len(duplicates) < 2:
            continue
        for item in duplicates:
            period_label = PERIOD_TYPE_LABELS.get(str(item.get("period_type") or ""), str(item.get("period_type") or "未注明期间"))
            basis_label = BASIS_LABELS.get(str(item.get("basis") or ""), str(item.get("basis") or "未注明口径"))
            if period_label not in item["label"] or basis_label not in item["label"]:
                raise SnapshotError(f"同名趋势指标未展示期间和口径：{company_id} / {base_label}")


def format_value(value: Any, unit: str) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "—"
    if unit == "%":
        return f"{number:,.2f}"
    if number.is_integer():
        return f"{number:,.0f}"
    return f"{number:,.2f}"


def comparison_html(value: dict[str, Any] | None) -> str:
    if not value or value.get("value") is None:
        return f'<span class="muted" title="{e((value or {}).get("reason_code") or "无正式比较")}">—</span>'
    number = float(value["value"])
    css_class = "positive" if number > 0 else "negative" if number < 0 else ""
    return f'<span class="{css_class}" title="程序按当前与可比期间计算">{number:+.2f}%</span>'


def render_financial_section(series: list[dict[str, Any]]) -> str:
    if not series:
        return '<section class="panel active" data-panel="financial"><div class="empty-state"><strong>暂无正式财务数据</strong></div></section>'
    options = "".join(f'<option value="{e(item["series_id"])}">{e(item["label"])}</option>' for item in series)
    tables = []
    for index, item in enumerate(series):
        rows = []
        for record in item["records"]:
            formula = ""
            if record.get("formula_id"):
                formula = f'<small>公式：{e(record.get("formula_id"))}</small>'
            rows.append(
                f'<tr><td><strong>{e(record.get("period"))}</strong><small>{e(record.get("period_start_date"))} → {e(record.get("period_end_date"))}</small></td>'
                f'<td><span class="metric-value">{e(format_value(record.get("value"), item["unit"]))}</span></td>'
                f'<td>{e(item["unit"])}</td><td>{comparison_html(record.get("qoq"))}</td><td>{comparison_html(record.get("yoy"))}</td>'
                f'<td><span class="nature">{e(record.get("data_nature"))}</span><small>{e(record.get("data_status"))}</small>{formula}</td>'
                f'<td><code>{e(record.get("record_id"))}</code></td></tr>'
            )
        definitions = "".join(
            f'<div><strong>{e(definition["name"])}</strong><code>{e(definition["metric_definition_id"])}</code>'
            f'<p>{e(definition["description"])}</p><small>{e(definition["scope_relationship"])} · {e(definition["verification_status"])}</small></div>'
            for definition in item["definitions"]
        )
        tables.append(
            f'<details class="metric-table" {"open" if index == 0 else ""}><summary><span>{e(item["label"])}</span>'
            f'<small>{len(item["records"])} 期 · {e(item["unit"])} · {e(item["basis"])}</small></summary>'
            '<div class="table-scroll"><table><thead><tr><th>财报期</th><th>数值</th><th>单位</th><th>环比</th><th>同比</th><th>性质 / 状态</th><th>数据记录编号</th></tr></thead>'
            f'<tbody>{"".join(rows)}</tbody></table></div><div class="definition-box"><p class="eyebrow">指标定义</p>{definitions}</div></details>'
        )
    return (
        '<section class="panel active" data-panel="financial"><div class="section-heading compact"><div><p class="eyebrow">历史财务数据</p>'
        '<h2>财务与运营趋势</h2></div><p>每个指标最多展示最近 8 个可比期间 · 缺失值不连线</p></div>'
        f'<div class="chart-shell"><div class="chart-toolbar"><label>趋势指标<select id="chart-series">{options}</select></label><div id="chart-unit"></div></div>'
        '<canvas id="trend-chart" height="300" aria-label="财务趋势图"></canvas><div id="chart-legend"></div></div>'
        f'<div class="metric-tables">{"".join(tables)}</div></section>'
    )


def render_coverage_section(series: list[dict[str, Any]]) -> str:
    periods = {record.get("period") for item in series for record in item["records"] if record.get("period")}
    records = [record for item in series for record in item["records"]]
    calculated = sum(record.get("data_nature") == "程序计算" for record in records)
    direct = len(records) - calculated
    rows = "".join(
        f'<tr><td><strong>{e(item["label"])}</strong></td><td>{len(item["records"])}</td><td>{e(item["period_type"])}</td>'
        f'<td>{e(item["unit"])}</td><td>{e(item["basis"])}</td></tr>'
        for item in series
    )
    return (
        '<section class="panel" data-panel="coverage"><div class="section-heading compact"><div><p class="eyebrow">正式数据覆盖</p><h2>数据覆盖</h2></div>'
        '<p>当前页面实际可浏览的数据范围</p></div><div class="coverage-cards">'
        f'<div><small>财报期间</small><strong>{len(periods)}</strong></div><div><small>指标序列</small><strong>{len(series)}</strong></div>'
        f'<div><small>公司披露记录</small><strong>{direct}</strong></div><div><small>程序计算记录</small><strong>{calculated}</strong></div></div>'
        '<div class="table-scroll"><table><thead><tr><th>指标 / 维度</th><th>已有期间</th><th>期间类型</th><th>单位</th><th>口径</th></tr></thead>'
        f'<tbody>{rows}</tbody></table></div></section>'
    )


def render_company_page(company: dict[str, Any], series: list[dict[str, Any]]) -> str:
    all_records = [record for item in series for record in item["records"]]
    latest = max(all_records, key=lambda row: (str(row.get("period_end_date") or ""), str(row.get("period") or "")))
    period_count = len({record.get("period") for record in all_records if record.get("period")})
    tabs = (
        ("financial", "财务与运营", len(series)),
        ("coverage", "数据覆盖", period_count),
    )
    tab_html = "".join(
        f'<button class="tab {"active" if key == "financial" else ""}" data-tab="{key}">{label}<b>{count}</b></button>'
        for key, label, count in tabs
    )
    body = (
        '<a class="back-link" href="/">← 返回公司列表</a><section class="company-hero"><div>'
        f'<p class="eyebrow">{e(company.get("市场") or "公司财报数据")}</p><h1>{e(company.get("公司"))}</h1>'
        f'<p>{e(company.get("股票代码") or company.get("company_id"))} · business / geography / product 独立口径</p></div></section>'
        '<section class="company-overview">'
        f'<span><small>最新财报期</small><strong>{e(latest.get("period"))}</strong></span>'
        f'<span><small>历史期间</small><strong>{period_count}</strong></span><span><small>数据指标</small><strong>{len(series)}</strong></span>'
        '<span class="internal-status">只读数据</span></section>'
        f'<nav class="tabs" aria-label="公司财报数据模块">{tab_html}</nav>'
        f'{render_financial_section(series)}{render_coverage_section(series)}'
    )
    if any(term in body for term in BLOCKED_BROWSER_TERMS):
        raise SnapshotError(f"公司数据页面仍包含研究模块：{company.get('company_id')}")
    return body


def normalize_calendar() -> list[dict[str, Any]]:
    rows = read_sheet("01_财报日历/财报日历.xlsx", "财报日历")
    normalized = []
    for row in rows:
        if not row.get("company_id"):
            continue
        status = str(row.get("状态") or "待核实")
        actual = row.get("实际披露北京时间")
        planned = row.get("官方计划披露北京时间")
        appointment = row.get("官方预约日期")
        estimate = row.get("第三方预计日期")
        released = bool(actual) or status == "已发布"
        normalized.append({
            "company_id": row.get("company_id"),
            "company": row.get("公司"),
            "ticker": row.get("股票代码"),
            "market": row.get("市场"),
            "period": row.get("财报期"),
            "report_type": row.get("财报类型"),
            "status": status,
            "released": released,
            "official_appointment_date": appointment,
            "planned_release_original": row.get("官方计划披露原始时间"),
            "planned_release_timezone": row.get("官方计划披露原始时区"),
            "planned_release_beijing": planned,
            "actual_release_original": row.get("实际披露原始时间"),
            "actual_release_timezone": row.get("实际披露原始时区"),
            "actual_release_beijing": actual,
            "estimated_date": estimate,
            "call_time_beijing": row.get("电话会北京时间"),
            "source_name": row.get("主要信源"),
            "source_url": row.get("信源链接"),
            "last_checked_at": row.get("最后核查时间"),
            "sort_at": actual or planned or appointment or estimate or "9999-12-31T00:00:00",
        })
    if not CALENDAR_SUPPLEMENTS_PATH.is_file():
        raise SnapshotError("前端财报预报补充文件缺失")
    supplements = json.loads(CALENDAR_SUPPLEMENTS_PATH.read_text(encoding="utf-8"))
    if not isinstance(supplements, list):
        raise SnapshotError("前端财报预报补充文件格式无效")
    seen_events = {(str(row.get("company_id")), str(row.get("period"))) for row in normalized}
    for row in supplements:
        required = ("event_id", "company_id", "company", "period", "official_appointment_date", "source_url", "event_basis")
        if not isinstance(row, dict) or any(not row.get(key) for key in required):
            raise SnapshotError("前端财报预报补充事件字段不完整")
        event_key = (str(row["company_id"]), str(row["period"]))
        if event_key in seen_events:
            continue
        if not str(row["source_url"]).startswith("https://"):
            raise SnapshotError(f"前端财报预报补充事件来源无效：{row['event_id']}")
        appointment = str(row["official_appointment_date"])
        normalized.append({
            "company_id": row["company_id"],
            "company": row["company"],
            "ticker": row.get("ticker"),
            "market": row.get("market"),
            "period": row["period"],
            "report_type": row.get("report_type") or "财报",
            "status": row.get("status") or "已确认",
            "released": False,
            "official_appointment_date": appointment,
            "planned_release_original": None,
            "planned_release_timezone": None,
            "planned_release_beijing": None,
            "actual_release_original": None,
            "actual_release_timezone": None,
            "actual_release_beijing": None,
            "estimated_date": None,
            "call_time_beijing": None,
            "source_name": row.get("source_name") or "公司公告",
            "source_url": row["source_url"],
            "last_checked_at": row.get("last_checked_at"),
            "event_basis": row["event_basis"],
            "frontend_supplement": True,
            "sort_at": appointment,
        })
        seen_events.add(event_key)
    return sorted(normalized, key=lambda row: (row["released"], row["sort_at"], row.get("company") or ""))


def load_calendar_coverage_reviews() -> list[dict[str, Any]]:
    try:
        payload = json.loads(CALENDAR_COVERAGE_REVIEWS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SnapshotError("财报预报覆盖复核文件缺失或格式无效") from exc
    if not isinstance(payload, list):
        raise SnapshotError("财报预报覆盖复核文件必须是数组")
    return payload


def event_date(row: dict[str, Any]) -> date | None:
    value = (
        row.get("planned_release_beijing")
        or row.get("official_appointment_date")
        or row.get("estimated_date")
    )
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError as exc:
        raise SnapshotError(f"财报预报事件日期无效：{row.get('company_id')} / {row.get('period')}") from exc


def verify_calendar_coverage(company_ids: set[str], calendar: list[dict[str, Any]]) -> None:
    today = date.today()
    upcoming_company_ids = set()
    for row in calendar:
        if row.get("released"):
            continue
        scheduled_date = event_date(row)
        if scheduled_date and scheduled_date < today:
            raise SnapshotError(
                f"未发布事件日期已过期，必须更新状态：{row.get('company_id')} / {row.get('period')} / {scheduled_date}"
            )
        if scheduled_date:
            upcoming_company_ids.add(str(row.get("company_id") or ""))

    active_reviews = set()
    seen_reviews = set()
    for row in load_calendar_coverage_reviews():
        required = ("company_id", "company", "status", "reviewed_at", "valid_until", "source_url", "note")
        if not isinstance(row, dict) or any(not row.get(key) for key in required):
            raise SnapshotError("财报预报覆盖复核记录字段不完整")
        company_id = str(row["company_id"])
        if company_id in seen_reviews:
            raise SnapshotError(f"财报预报覆盖复核记录重复：{company_id}")
        seen_reviews.add(company_id)
        if row["status"] != "needs_m1_review" or not str(row["source_url"]).startswith("https://"):
            raise SnapshotError(f"财报预报覆盖复核记录状态或来源无效：{company_id}")
        try:
            reviewed_at = date.fromisoformat(str(row["reviewed_at"]))
            valid_until = date.fromisoformat(str(row["valid_until"]))
        except ValueError as exc:
            raise SnapshotError(f"财报预报覆盖复核日期无效：{company_id}") from exc
        if reviewed_at > today or valid_until < reviewed_at or (valid_until - reviewed_at).days > 7:
            raise SnapshotError(f"财报预报覆盖复核窗口无效：{company_id}")
        if valid_until < today:
            raise SnapshotError(f"财报预报覆盖复核已过期：{company_id} / {valid_until}")
        active_reviews.add(company_id)

    uncovered = sorted(company_ids - upcoming_company_ids - active_reviews)
    if uncovered:
        raise SnapshotError(f"关注公司缺少下一次财报事件或有效复核记录：{', '.join(uncovered)}")


def build_snapshot() -> dict[str, Any]:
    manifest = source_manifest()
    frontend_manifest = frontend_input_manifest()
    calendar = normalize_calendar()
    registry = load_registry()
    companies = {
        str(row["company_id"]): row
        for row in read_sheet("00_系统/关注公司.xlsx", "关注公司")
        if row.get("company_id") and str(row.get("是否关注") or "") == "是"
    }
    financial_rows = [
        row for row in read_sheet("04_公司数据库/财报历史数据库.xlsx", "历史数据")
        if row.get("record_id") and current_row(row)
    ]
    rows_by_company: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in financial_rows:
        if row.get("company_id"):
            rows_by_company[str(row["company_id"])].append(row)

    company_pages: dict[str, dict[str, Any]] = {}
    for company_id, company in companies.items():
        series = build_series(rows_by_company.get(company_id, []), registry)
        if not series:
            continue
        verify_series_labels(company_id, series)
        all_records = [record for item in series for record in item["records"]]
        latest = max(all_records, key=lambda row: (str(row.get("period_end_date") or ""), str(row.get("period") or "")))
        periods = {record.get("period") for record in all_records if record.get("period")}
        company_pages[company_id] = {
            "company_id": company_id,
            "company": company.get("公司"),
            "ticker": company.get("股票代码"),
            "market": company.get("市场"),
            "target_period": latest.get("period"),
            "financial_period_count": len(periods),
            "metric_count": len(series),
            "html": render_company_page(company, series),
            "page_data": {"financial_series": series},
        }
    if not company_pages:
        raise SnapshotError("数据平台没有可展示的公司财务数据")
    verify_calendar_coverage(set(company_pages), calendar)

    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    snapshot_seed = {
        "schema": SNAPSHOT_SCHEMA,
        "company_ids": sorted(company_pages),
        "source_manifest": manifest,
        "frontend_input_manifest": frontend_manifest,
    }
    payload: dict[str, Any] = {
        "schema_version": SNAPSHOT_SCHEMA,
        "snapshot_id": "dps_" + sha256_bytes(stable_json(snapshot_seed).encode("utf-8"))[:24],
        "generated_at": generated_at,
        "derived": True,
        "authoritative": False,
        "production_mutation": False,
        "truth_source": "Local ResearchOS production remains the sole truth source and writer",
        "access_intent": "public_financial_data_platform",
        "source_manifest": manifest,
        "frontend_input_manifest": frontend_manifest,
        "summary": {
            "company_count": len(company_pages),
            "calendar_event_count": len(calendar),
            "upcoming_event_count": sum(not row["released"] for row in calendar),
            "released_event_count": sum(row["released"] for row in calendar),
        },
        "earnings_calendar": calendar,
        "company_pages": company_pages,
    }
    payload["snapshot_content_sha256"] = sha256_bytes(stable_json(payload).encode("utf-8"))
    return payload


def verify_snapshot(payload: dict[str, Any]) -> None:
    if payload.get("schema_version") != SNAPSHOT_SCHEMA:
        raise SnapshotError("data platform snapshot schema 无效")
    if payload.get("derived") is not True or payload.get("authoritative") is not False:
        raise SnapshotError("snapshot 派生属性无效")
    if payload.get("production_mutation") is not False:
        raise SnapshotError("snapshot 错误标记了 production mutation")
    if payload.get("access_intent") != "public_financial_data_platform":
        raise SnapshotError("snapshot access intent 无效")
    verify_file_manifest(payload.get("source_manifest"), RESEARCHOS_ROOT, "正式来源")
    verify_file_manifest(payload.get("frontend_input_manifest"), PROJECT_DIR, "前端规则输入")
    expected = payload.get("snapshot_content_sha256")
    content = dict(payload)
    content.pop("snapshot_content_sha256", None)
    if expected != sha256_bytes(stable_json(content).encode("utf-8")):
        raise SnapshotError("snapshot content hash 不匹配")
    if not isinstance(payload.get("earnings_calendar"), list):
        raise SnapshotError("snapshot 财报预报数据无效")
    pages = payload.get("company_pages")
    if not isinstance(pages, dict) or not pages:
        raise SnapshotError("snapshot 没有公司财务数据页面")
    verify_calendar_coverage(set(pages), payload["earnings_calendar"])
    for company_id, page in pages.items():
        if not page.get("html") or not page.get("page_data", {}).get("financial_series"):
            raise SnapshotError(f"snapshot 公司数据页面不完整：{company_id}")
        if any(term in page["html"] for term in BLOCKED_BROWSER_TERMS):
            raise SnapshotError(f"snapshot 公司页面包含已删除的研究模块：{company_id}")
        verify_series_labels(company_id, page["page_data"]["financial_series"])


def export() -> dict[str, Any]:
    payload = build_snapshot()
    verify_snapshot(payload)
    atomic_write(
        SNAPSHOT_PATH,
        (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
    )
    for asset_name in ("snapshot-styles.css", "snapshot-polish.css", "snapshot-app.js"):
        if not (PUBLIC_DIR / asset_name).is_file():
            raise SnapshotError(f"数据展示资产缺失：{asset_name}")
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="生成财报预报与财务数据公开只读 Snapshot")
    parser.add_argument("--verify", action="store_true", help="验证已生成 snapshot")
    args = parser.parse_args()
    try:
        payload = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8")) if args.verify else export()
        verify_snapshot(payload)
        print(json.dumps({
            "status": "PASS",
            "snapshot_id": payload["snapshot_id"],
            "company_count": len(payload["company_pages"]),
            "calendar_event_count": len(payload["earnings_calendar"]),
            "production_mutation": payload["production_mutation"],
        }, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(json.dumps({"status": "FAIL_CLOSED", "error": str(exc)}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
